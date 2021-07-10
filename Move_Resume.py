import os

from openpyxl import load_workbook
from shutil import move

#excel简历信息库的文件名，需根据当天的文件名更改
exl_file='D:\Project\技术分享\MoveResume\附件2：xx校招信息库-20210420-306份-累计推荐2277份.xlsx'
#需要检索的简历存放路径，更改路径时注意不要误删路径两头的单引号'和最前面的字母r
path=r'\\ssawd404\部门建设\21 招聘\校园招聘简历\xx春季校园招聘简历推送-20210420'
#筛选后的简历存放路径，更改路径时注意不要误删路径两头的单引号'和最前面的字母r
dst_path=r'D:\Project\技术分享\MoveResume\筛后简历'

i = 0 # i记录找到的简历数量


def main():
    """
    根据简历库的筛选结果从公共盘中剪切出筛到的候选人简历供组内同事面试
    时间：2021/4/18
    """

    #S1：读excel信息库找到筛选出的名字
    wb=load_workbook(filename=exl_file)
    ws=wb['待提取候选人']
    #对名字所在的列进行遍历
    for cell in ws["B"]:
        #S2:去文件夹中剪切想要的简历
        move_resume(str(cell.value))
    if i>0:
        print("已有",i,"份简历被您从\n", path, "\n", "剪切至\n", dst_path)
        print(" Congratulations!")
    else:
        print(" 没有简历被剪走")
        print(" 请检查是否已剪切完毕，或程序中的文件和路径名是否正确")

def move_resume(name):
    global i
    for root,lists,files in os.walk(path):
        for file in files:
            if (name in file) & (file[0]!="."): #去掉隐藏文件
                i=i+1
                src_path=os.path.join(root,file) #该简历所在路径
                try:
                    move(src_path, dst_path)
                except: #防止文件正在使用，无法剪切
                    pass

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()